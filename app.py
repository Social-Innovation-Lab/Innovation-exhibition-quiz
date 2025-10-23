import sqlite3
import csv
import secrets
from datetime import datetime
from flask import Flask, render_template, request, redirect, session, jsonify, make_response, abort
from functools import wraps
import os
import hashlib
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = os.environ.get('SESSION_SECRET', secrets.token_hex(32))
# Configure session for iframe/preview environment
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Allows cookies in iframe redirects
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = 3600  # 1 hour session

DATABASE = 'quiz.db'
ADMIN_PIN = os.environ.get('ADMIN_PIN', '2025')
EXCEL_FILE = 'quiz_results.xlsx'

def generate_csrf_token():
    """Generate CSRF token for session"""
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']

def verify_csrf_token(token):
    """Verify CSRF token matches session"""
    return token and session.get('csrf_token') == token

def admin_required(f):
    """Decorator to require admin authentication"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_authenticated'):
            return redirect('/admin/login')
        return f(*args, **kwargs)
    return decorated_function

@app.context_processor
def inject_csrf_token():
    """Make CSRF token available to all templates"""
    return dict(csrf_token=generate_csrf_token)

def get_db():
    """Get database connection with WAL mode"""
    conn = sqlite3.connect(DATABASE, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    return conn

def init_db():
    """Initialize database schema and import questions"""
    conn = get_db()
    
    # Drop old tables to start fresh with new schema
    conn.executescript('''
        DROP TABLE IF EXISTS responses;
        DROP TABLE IF EXISTS attempts;
        DROP TABLE IF EXISTS rotation_queue;
        DROP TABLE IF EXISTS questions;
    ''')
    
    # Create tables with new schema
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            programme_code TEXT NOT NULL,
            programme_name TEXT NOT NULL,
            difficulty TEXT NOT NULL,
            weight REAL NOT NULL,
            question TEXT NOT NULL,
            option_A TEXT NOT NULL,
            option_B TEXT NOT NULL,
            option_C TEXT NOT NULL,
            option_D TEXT NOT NULL,
            answer TEXT NOT NULL,
            answer_text TEXT NOT NULL
        );
        
        CREATE TABLE IF NOT EXISTS participants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            pin TEXT NOT NULL,
            phone TEXT NOT NULL,
            consent INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        
        CREATE TABLE IF NOT EXISTS attempts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            participant_id INTEGER NOT NULL,
            score INTEGER NOT NULL,
            total_questions INTEGER DEFAULT 10,
            percent REAL NOT NULL,
            is_winner INTEGER DEFAULT 0,
            gift_given INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (participant_id) REFERENCES participants(id)
        );
        
        CREATE TABLE IF NOT EXISTS responses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            attempt_id INTEGER NOT NULL,
            question_id INTEGER NOT NULL,
            selected_answer TEXT NOT NULL,
            is_correct INTEGER NOT NULL,
            FOREIGN KEY (attempt_id) REFERENCES attempts(id),
            FOREIGN KEY (question_id) REFERENCES questions(id)
        );
        
        CREATE INDEX IF NOT EXISTS idx_questions_difficulty ON questions(difficulty);
        CREATE INDEX IF NOT EXISTS idx_attempts_winner ON attempts(is_winner);
    ''')
    
    # Import questions
    print("Importing questions from CSV...")
    import_questions()
    
    conn.commit()
    conn.close()
    print("Database initialized successfully!")

def capitalize_first_letter(text):
    """Capitalize the first letter of a string if it's not already capitalized"""
    if not text:
        return text
    return text[0].upper() + text[1:] if text else text

def import_questions():
    """Import questions from CSV file with actual programme names"""
    import re
    conn = get_db()
    csv_path = 'attached_assets/Untitled spreadsheet - QuestionBank_1761118191656.csv'
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Keep the actual question text with programme names
            question_text = row['question']
            
            # Clean up extra spaces and punctuation
            question_text = re.sub(r'\s+', ' ', question_text).strip()
            question_text = re.sub(r'\s+([?.!,])', r'\1', question_text)  # Fix spacing before punctuation
            
            # Capitalize first letter
            question_text = capitalize_first_letter(question_text)
            
            # Capitalize options
            option_a = capitalize_first_letter(row['option_A'].strip())
            option_b = capitalize_first_letter(row['option_B'].strip())
            option_c = capitalize_first_letter(row['option_C'].strip())
            option_d = capitalize_first_letter(row['option_D'].strip())
            
            conn.execute('''
                INSERT INTO questions (programme_code, programme_name, difficulty, weight, question, 
                                      option_A, option_B, option_C, option_D, answer, answer_text)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (row['programme_code'], row['programme_name'], row['difficulty'], 
                  float(row['weight']), question_text,
                  option_a, option_b, option_c, option_d,
                  row['answer_letter'], row['answer_text']))
    
    conn.commit()
    print(f"Imported questions from new question bank")

def select_weighted_random_questions(num_questions=10):
    """Select random questions with difficulty-based weighting
    
    Target distribution for 10 questions:
    - Easy (weight 1): 3 questions (30%)
    - Medium (weight 1.5): 3 questions (30%)
    - Hard (weight 2): 4 questions (40%)
    """
    import random
    conn = get_db()
    
    # Target distribution
    target_easy = 3
    target_medium = 3
    target_hard = 4
    
    selected_questions = []
    
    # Get Easy questions
    easy_questions = conn.execute(
        'SELECT * FROM questions WHERE difficulty = "Easy" ORDER BY RANDOM() LIMIT ?',
        (target_easy,)
    ).fetchall()
    selected_questions.extend([dict(q) for q in easy_questions])
    
    # Get Medium questions
    medium_questions = conn.execute(
        'SELECT * FROM questions WHERE difficulty = "Medium" ORDER BY RANDOM() LIMIT ?',
        (target_medium,)
    ).fetchall()
    selected_questions.extend([dict(q) for q in medium_questions])
    
    # Get Hard questions
    hard_questions = conn.execute(
        'SELECT * FROM questions WHERE difficulty = "Hard" ORDER BY RANDOM() LIMIT ?',
        (target_hard,)
    ).fetchall()
    selected_questions.extend([dict(q) for q in hard_questions])
    
    # Shuffle the selected questions so difficulty levels are mixed
    random.shuffle(selected_questions)
    
    return selected_questions

def export_to_excel(participant_id, score, percent, is_winner, total_questions=10):
    """Export quiz result to Excel file"""
    try:
        conn = get_db()
        participant = conn.execute(
            'SELECT name, pin, phone FROM participants WHERE id = ?',
            (participant_id,)
        ).fetchone()
        
        if not participant:
            return
        
        # Create or load workbook
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Quiz Results"
            
            # Create header row
            headers = ['Name', 'PIN', 'Phone', 'Score', 'Percentage', 'Winner', 'Date & Time']
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="E31837", end_color="E31837", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Mask phone number (show last 4 digits only)
        phone = participant['phone']
        masked_phone = f"****{phone[-4:]}" if len(phone) >= 4 else phone
        
        # Add data row
        row_data = [
            participant['name'],
            participant['pin'],
            masked_phone,
            f"{score}/{total_questions}",
            f"{percent:.2f}%",
            "Yes" if is_winner else "No",
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
        ws.append(row_data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(EXCEL_FILE)
        print(f"Exported result to {EXCEL_FILE}")
        
    except Exception as e:
        print(f"Error exporting to Excel: {e}")

@app.route('/')
def index():
    """Landing page with sign-in form"""
    session.clear()
    return render_template('index.html')

@app.route('/start', methods=['POST'])
def start():
    """Handle participant sign-in and start quiz"""
    name = request.form.get('name', '').strip()
    pin = request.form.get('pin', '').strip()
    phone = request.form.get('phone', '').strip()
    
    # Validate all required fields and PIN format (exactly 6 digits)
    if not all([name, pin, phone]) or not (pin.isdigit() and len(pin) == 6):
        print(f"Validation failed - name:{name}, pin:{pin}, phone:{phone}")
        return redirect('/')
    
    # Create participant record
    conn = get_db()
    cursor = conn.execute(
        'INSERT INTO participants (name, pin, phone, consent) VALUES (?, ?, ?, ?)',
        (name, pin, phone, 1)
    )
    participant_id = cursor.lastrowid
    conn.commit()
    
    print(f"Created participant: participant_id={participant_id}, name={name}")
    
    # Get 10 random questions with weighted distribution
    questions = select_weighted_random_questions(10)
    
    # Format for template
    for q in questions:
        q['options'] = [q['option_A'], q['option_B'], q['option_C'], q['option_D']]
    
    # Store the participant_id and question IDs to pass to template
    question_ids = [q['id'] for q in questions]
    
    print(f"Selected {len(questions)} questions for quiz")
    
    # Render quiz page directly (no redirect to avoid cookie issues)
    return render_template('quiz.html', 
                         items=questions, 
                         participant_id=participant_id,
                         question_ids=question_ids,
                         participant_name=name)

@app.route('/submit', methods=['POST'])
def submit():
    """Grade quiz and save results"""
    # Get participant data from form instead of session (to avoid cookie issues)
    participant_id = request.form.get('participant_id', '').strip()
    question_ids_str = request.form.get('question_ids', '').strip()
    
    if not participant_id or not question_ids_str:
        return redirect('/')
    
    participant_id = int(participant_id)
    question_ids = [int(qid) for qid in question_ids_str.split(',') if qid]
    
    conn = get_db()
    
    # Total questions is now 10
    total_questions = len(question_ids)
    
    # Create attempt record
    cursor = conn.execute(
        'INSERT INTO attempts (participant_id, score, total_questions, percent, is_winner) VALUES (?, 0, ?, 0, 0)',
        (participant_id, total_questions)
    )
    attempt_id = cursor.lastrowid
    
    # Grade responses
    score = 0
    for qid in question_ids:
        selected = request.form.get(str(qid), '').strip()
        
        # Get correct answer
        question = conn.execute('SELECT answer FROM questions WHERE id = ?', (qid,)).fetchone()
        is_correct = 1 if (question and selected == question['answer']) else 0
        score += is_correct
        
        # Save response
        conn.execute(
            'INSERT INTO responses (attempt_id, question_id, selected_answer, is_correct) VALUES (?, ?, ?, ?)',
            (attempt_id, qid, selected, is_correct)
        )
    
    # Calculate percentage and winner status (70% = 7/10)
    percent = (score / total_questions) * 100
    is_winner = 1 if score >= 7 else 0
    
    # Update attempt
    conn.execute(
        'UPDATE attempts SET score = ?, percent = ?, is_winner = ? WHERE id = ?',
        (score, percent, is_winner, attempt_id)
    )
    
    conn.commit()
    
    print(f"Quiz graded: score={score}/{total_questions}, percent={percent:.2f}%, winner={is_winner}")
    
    # Export result to Excel file
    export_to_excel(participant_id, score, percent, is_winner, total_questions)
    
    # Get participant name for display
    participant = conn.execute(
        'SELECT name FROM participants WHERE id = ?', (participant_id,)
    ).fetchone()
    
    # Render result page directly (no redirect to avoid cookie issues)
    data = {
        'name': participant['name'] if participant else 'Participant',
        'score': score,
        'total': total_questions,
        'percent': round(percent, 2),
        'is_winner': is_winner
    }
    
    return render_template('result.html', data=data)

@app.route('/result')
def result():
    """Display quiz results"""
    if 'attempt_id' not in session:
        return redirect('/')
    
    attempt_id = session['attempt_id']
    conn = get_db()
    
    attempt = conn.execute(
        'SELECT * FROM attempts WHERE id = ?', (attempt_id,)
    ).fetchone()
    
    if not attempt:
        return redirect('/')
    
    data = {
        'score': attempt['score'],
        'percent': attempt['percent'],
        'is_winner': attempt['is_winner']
    }
    
    return render_template('result.html', data=data)

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page"""
    if request.method == 'POST':
        csrf_token = request.form.get('csrf_token')
        if not verify_csrf_token(csrf_token):
            abort(403)
        
        pin = request.form.get('admin_pin', '')
        if pin == ADMIN_PIN:
            session['admin_authenticated'] = True
            return redirect('/admin')
        else:
            return render_template('admin_login.html', error='Invalid PIN')
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    """Admin logout"""
    session.pop('admin_authenticated', None)
    return redirect('/')

@app.route('/admin')
@admin_required
def admin():
    """Admin dashboard with winner list"""
    conn = get_db()
    
    # Get all attempts with participant info
    attempts = conn.execute('''
        SELECT a.*, p.name, p.pin, p.phone
        FROM attempts a
        JOIN participants p ON a.participant_id = p.id
        ORDER BY a.created_at DESC
    ''').fetchall()
    
    # Get statistics
    total_attempts = len(attempts)
    total_winners = conn.execute('SELECT COUNT(*) FROM attempts WHERE is_winner = 1').fetchone()[0]
    avg_score = conn.execute('SELECT AVG(score) FROM attempts').fetchone()[0] or 0
    
    stats = {
        'total_attempts': total_attempts,
        'total_winners': total_winners,
        'avg_score': round(avg_score, 1)
    }
    
    return render_template('admin.html', attempts=attempts, stats=stats)

@app.route('/admin/mark_gift/<int:attempt_id>', methods=['POST'])
@admin_required
def mark_gift(attempt_id):
    """Mark gift as given"""
    csrf_token = request.form.get('csrf_token')
    if not verify_csrf_token(csrf_token):
        abort(403)
    
    conn = get_db()
    conn.execute('UPDATE attempts SET gift_given = 1 WHERE id = ?', (attempt_id,))
    conn.commit()
    return redirect('/admin')

@app.route('/admin/export')
@admin_required
def export_csv():
    """Export winners to CSV"""
    conn = get_db()
    
    winners = conn.execute('''
        SELECT p.name, p.pin, p.phone, a.score, a.percent, a.created_at, a.gift_given
        FROM attempts a
        JOIN participants p ON a.participant_id = p.id
        WHERE a.is_winner = 1
        ORDER BY a.created_at DESC
    ''').fetchall()
    
    # Create CSV
    output = "Name,PIN,Phone,Score,Percentage,Date,Gift Given\n"
    for w in winners:
        # Mask phone number (show only last 4 digits)
        masked_phone = '****' + w['phone'][-4:] if len(w['phone']) >= 4 else w['phone']
        gift_status = 'Yes' if w['gift_given'] else 'No'
        output += f'"{w["name"]}",{w["pin"]},{masked_phone},{w["score"]},{w["percent"]:.2f}%,{w["created_at"]},{gift_status}\n'
    
    response = make_response(output)
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = 'attachment; filename=winners.csv'
    return response

@app.route('/manifest.json')
def manifest():
    """PWA manifest"""
    return jsonify({
        "name": "BRAC Exhibition Quiz",
        "short_name": "BRAC Quiz",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#ffffff",
        "theme_color": "#e31837",
        "icons": [
            {"src": "/static/icons/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/icons/icon-512.png", "sizes": "512x512", "type": "image/png"}
        ]
    })

# Initialize database on startup (before app runs)
if not os.path.exists(DATABASE):
    init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
